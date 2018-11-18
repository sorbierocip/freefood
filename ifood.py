import time
import requests
from bs4 import BeautifulSoup
import urllib3
from openpyxl import Workbook
import  os

urllib3.disable_warnings()

if __name__ == '__main__':
    cep = input("Indique o CEP(somente numeros): ")
    number = input("Indique o número do endereço:")
    value = float(input("Indique o valor do cupom (10,00):").replace(',','.'))

    proxiesJson = None
    #proxiesJson ={'ip':'127.0.0.1', 'port':8888}
    sess = requests.session()
    if proxiesJson:
        proxies = {
            'http': 'http://' + proxiesJson['ip'] + ':' + str(proxiesJson['port']),
            'https': 'https://' + proxiesJson['ip'] + ':' + str(proxiesJson['port'])
        }
        sess.proxies.update(proxies)
    sess.verify = False
    sess.get('https://www.ifood.com.br/')

    location = sess.post('https://www.ifood.com.br/location/by-zip-code', data={'zipCode':cep,
                                                                     'address':'',
                                                                     'streetNumber':number
                                                                      },
              headers={'X-Requested-With': 'XMLHttpRequest'}).json()['Records'][0]


    sess.post('https://www.ifood.com.br/lista-restaurantes', data={
        'location.locationId':location['locationId'],
        'location.lat': location['lat'],
        'location.lon': location['lon'],
        'location.zipCode': location['zipCode'],
        'location.district': location['district'],
        'location.city': location['city'],
        'location.state': location['state'],
        'location.country': location['country'],
        'location.address': location['address'],
        'location.requireCompl': location['requireCompl'],
        'compl': '',
        'alias': '',
        'reference': '',
        'streetNumber': number
    })


    page = 1
    noStop = True
    stores = []

    while noStop:
        listStorePage = sess.post('https://www.ifood.com.br/lista-restaurantes/filtro',
                  data={'page':page,
                        'locationId':location['locationId'],
                        'city':'',
                        'state':'',
                        'ordenacao':'0',
                       'free-delivery':'on'})

        soup = BeautifulSoup(listStorePage.content, 'html.parser')

        storesLink = soup.find_all('a',class_='restaurant-card-link')
        for storeLink in storesLink:
            if 'open' in storeLink('article')[0]['class']:
                stores.append({'rid':storeLink['data-rid'],
                               'url': 'https://www.ifood.com.br/' + storeLink['href'],
                               'name':storeLink['data-name']})
            else:
                noStop=False
        page = page + 1


    possibleStores = []
    for item in stores:
        storePage = sess.get('https://www.ifood.com.br/delivery/refresh-cart?rid=' +item['rid']+'&_='+str(int(round(time.time() * 1000))))

        if 'Pedido M&iacute;nimo' not in storePage.text:
            item['pedidoMinimo'] = float(0)
            possibleStores.append(item)
        else:
            soup = BeautifulSoup(storePage.content, 'html.parser')
            minium = float(soup.find('div', class_='clearfix minimum')('span')[1].text.replace('R$','').strip().replace(',','.'))
            item['pedidoMinimo'] = minium
            if minium <= value:
                possibleStores.append(item)

    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Nome'
    ws['B1'] = 'Link'
    ws['C1'] = 'Pedido Minimo'

    for idx,item in enumerate(possibleStores):
        ws['A' + str(idx + 2)] = item['name']
        ws['B' + str(idx + 2)] = item['url']
        ws['C' + str(idx + 2)] = item['pedidoMinimo']

    wb.save("ifood.xlsx")

    os.system("start " + "ifood.xlsx")

